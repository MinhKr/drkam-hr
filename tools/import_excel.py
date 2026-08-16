# -*- coding: utf-8 -*-
"""
Nạp dữ liệu thật từ DATA UV DRKAM 2026.xlsx vào cơ sở dữ liệu.

  Xem trước, không ghi gì (nên chạy đầu tiên):
      python tools/import_excel.py

  Sinh file SQL để dán vào Supabase SQL Editor:
      python tools/import_excel.py --sql web/supabase/seed_data.sql

  Đẩy thẳng lên Supabase (cần NEXT_PUBLIC_SUPABASE_URL và SUPABASE_SERVICE_ROLE_KEY
  trong web/.env.local):
      python tools/import_excel.py --push

Chuyển đổi tự động khi nạp:
  - chuẩn hoá số điện thoại, giữ số 0 đầu
  - gộp giá trị danh mục trùng nghĩa (Face -> Facebook, VN Work -> Vietnamworks)
  - tách ngày/giờ phỏng vấn thành bản ghi riêng cho vòng 1 và vòng 2
  - ghép sheet onboard theo họ tên
  - đánh dấu hồ sơ nghi trùng theo SĐT/email để HR xem lại
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import pathlib
import re
import sys
import urllib.request
from collections import Counter, defaultdict

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

ROOT = pathlib.Path(__file__).resolve().parent.parent
XLSX = ROOT / "DATA UV DRKAM 2026.xlsx"
ENV = ROOT / "web" / ".env.local"

# Bảng nằm trong schema riêng vì dùng chung project Supabase với app khác
DB_SCHEMA = "tuyendung"

SOURCE_ALIASES = {
    "face": "Facebook",
    "vn work": "Vietnamworks",
    "vietnamwork": "Vietnamworks",
    "career link": "CareerLink",
    "careerlink": "CareerLink",
    "career viet": "CareerViet",
    "hunt a": "Hunt",
    "thereads": "Threads",
    "josbgo": "JobsGO",
    "linkedin": "LinkedIn",
    "lọc top cv": "TopCV (chủ động lọc)",
    "chị diệu linh giới thiệu": "Nội bộ giới thiệu",
    "trade cv": "TradeCV",
    "topcv": "TopCV",
    "facebook": "Facebook",
}

POSITION_ALIASES = {
    "mkt ai": "Marketing AI",
    "marketing ai": "Marketing AI",
    "digital mkt": "Digital Marketing",
    "tp mkt": "Trưởng phòng Marketing",
    "des": "Designer",
    "kt sàn": "Kế toán sàn",
    "kt kho": "Kế toán kho",
    "kt viên": "Kế toán viên",
    "kt kiêm hc": "Kế toán kiêm Hành chính",
}


# ----------------------------------------------------------------- tiện ích
def txt(v) -> str | None:
    if v is None:
        return None
    s = re.sub(r'"?&(amp;)?"?', "", str(v)) if '"&' in str(v) else str(v)
    s = re.sub(r"\s+", " ", s).strip()
    return s or None


def alias(v: str | None, table: dict) -> str | None:
    if not v:
        return None
    return table.get(v.lower().strip(), v)


def sdt(v) -> str | None:
    if v is None:
        return None
    d = re.sub(r"\D", "", str(v))
    if not d:
        return None
    if d.startswith("84") and len(d) >= 11:
        d = "0" + d[2:]
    elif not d.startswith("0") and len(d) == 9:
        d = "0" + d
    return d


def ngay(v) -> str | None:
    if v is None:
        return None
    if isinstance(v, (dt.datetime, dt.date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    for f in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return dt.datetime.strptime(s[:10], f).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def gio(v) -> str | None:
    if v is None:
        return None
    if isinstance(v, dt.time):
        return v.strftime("%H:%M:%S")
    if isinstance(v, dt.datetime):
        return v.strftime("%H:%M:%S")
    s = str(v).strip()
    m = re.match(r"^(\d{1,2})[:h](\d{2})", s)
    if m:
        return f"{int(m.group(1)):02d}:{m.group(2)}:00"
    return None


def yes(v) -> bool:
    return str(v).strip().lower() in {"yes", "y", "x", "true", "có", "co", "1"}


def doc_env() -> dict:
    if not ENV.exists():
        return {}
    out = {}
    for line in ENV.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        out[k.strip()] = v.strip().strip('"').strip("'")
    return out


# -------------------------------------------------------------------- đọc
def doc_ung_vien(wb):
    ws = wb["Data"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    header = [txt(h) for h in rows[0]]
    idx = {h: i for i, h in enumerate(header) if h}

    def g(row, ten):
        i = idx.get(ten)
        return row[i] if i is not None and i < len(row) else None

    ung_vien, phong_van = [], []
    for row in rows[1:]:
        ten = txt(g(row, "Họ và tên"))
        if not ten:
            continue

        uv = {
            "received_at": ngay(g(row, "Ngày nhận CV")) or dt.date.today().isoformat(),
            "full_name": ten,
            "gender": txt(g(row, "Giới tính")),
            "region": txt(g(row, "Khu vực Ứng tuyển")),
            "phone": sdt(g(row, "Số điện thoại")),
            "email": (txt(g(row, "Email")) or "").lower() or None,
            "cv_url": txt(g(row, "Link CV/Portfolio")),
            "position": alias(txt(g(row, "Vị trí ứng tuyển")), POSITION_ALIASES),
            "department": txt(g(row, "Phòng ban")),
            "level": txt(g(row, "Cấp bậc")),
            "source": alias(txt(g(row, "Nguồn CV")), SOURCE_ALIASES),
            "screener": txt(g(row, "Người sàng lọc CV")),
            "screening_note": txt(g(row, "KQ sàng lọc")),
            "status": txt(g(row, "Trạng thái CV")) or "Đang liên hệ",
            "hometown": txt(g(row, "Quê quán")),
            "experience": txt(g(row, "Kinh nghiệm")),
            "available_from": txt(g(row, "Thời gian onboard")),
            "expected_salary": txt(g(row, "Mức lương mong muốn")),
            "offer_status": txt(g(row, "Trạng thái offer")),
            "planned_onboard_date": ngay(g(row, "Ngày dự kiến onboard")),
            "actual_onboard_date": ngay(g(row, "Ngày thực tế onboard")),
        }
        ung_vien.append(uv)

        for vong, hau_to in ((1, "1"), (2, "2")):
            d = ngay(g(row, f"Ngày PV{hau_to}"))
            kq = txt(g(row, f"KQ PV{hau_to}"))
            nguoi = txt(g(row, f"Người PV{hau_to}"))
            if not (d or kq or nguoi):
                continue
            phong_van.append(
                {
                    "_ten": ten,
                    "_sdt": uv["phone"],
                    "round": vong,
                    "scheduled_date": d,
                    "scheduled_time": gio(g(row, f"Giờ PV{hau_to}")),
                    "mode": txt(g(row, f"Hình thức PV{hau_to}")),
                    "interviewers": [p.strip() for p in (nguoi or "").split(",") if p.strip()],
                    "result": kq,
                    "note": txt(g(row, f"Note PV{hau_to}")),
                    "result_email_sent": yes(g(row, "Trả kết quả pv qua Mail")) if vong == 1 else False,
                }
            )

    return ung_vien, phong_van


CHECKLIST_COT = [
    ("pre_group", 8), ("pre_csvc", 9),
    ("day_den", 12), ("day_ho_so", 13), ("day_thiet_bi", 14),
    ("kb_bhxh", 15), ("kb_tncn", 16), ("kb_mst", 17),
    ("ck_tai_nguyen", 18), ("ck_ho_so", 19), ("ck_cham_cong", 20),
    ("ck_dao_tao", 21), ("ck_boi_thuong", 22),
    ("dt_doanh_nghiep", 24), ("dt_san_pham", 25), ("dt_ai", 26), ("dt_tai_lieu_ai", 27),
]


def doc_onboard(wb):
    ws = wb["LỊCH ONBOARD UV"]
    out = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        ten = txt(row[3]) if len(row) > 3 else None
        if not ten:
            continue
        out.append(
            {
                "_ten": ten,
                "_sdt": sdt(row[5]) if len(row) > 5 else None,
                "onboard_date": ngay(row[2]),
                "office": txt(row[7]) if len(row) > 7 else None,
                "checklist": {k: yes(row[i]) for k, i in CHECKLIST_COT if i < len(row)},
                "assignee_pre": txt(row[10]) if len(row) > 10 else None,
                "pre_note": txt(row[11]) if len(row) > 11 else None,
                "assignee_docs": txt(row[23]) if len(row) > 23 else None,
                "assignee_training": txt(row[28]) if len(row) > 28 else None,
                "status": txt(row[29]) if len(row) > 29 else None,
                "review_7d_result": txt(row[31]) if len(row) > 31 else None,
                "review_1m_result": txt(row[33]) if len(row) > 33 else None,
                "review_2m_result": txt(row[35]) if len(row) > 35 else None,
                "note": txt(row[36]) if len(row) > 36 else None,
                "owner": txt(row[37]) if len(row) > 37 else None,
            }
        )
    return out


# ------------------------------------------------------------------ xuất
def q(v) -> str:
    if v is None:
        return "null"
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, (int, float)):
        return str(v)
    if isinstance(v, (list, dict)):
        return "'" + json.dumps(v, ensure_ascii=False).replace("'", "''") + "'"
    return "'" + str(v).replace("'", "''") + "'"


def sinh_sql(ung_vien, phong_van, onboard) -> str:
    L = [
        "-- Dữ liệu ứng viên — sinh từ DATA UV DRKAM 2026.xlsx",
        "-- Chạy SAU 0001_init.sql và 0002_seed_catalogs.sql",
        "begin;",
        "",
        "-- Tắt trigger trong lúc nạp: giữ nguyên trạng thái lịch sử của từng ứng viên",
        "-- (nếu để bật, trigger sẽ tính lại trạng thái từ kết quả PV và ghi đè dữ liệu cũ)",
        "-- đồng thời không đổ hàng nghìn dòng vào nhật ký thay đổi.",
        "alter table tuyendung.interviews  disable trigger interviews_sync_status;",
        "alter table tuyendung.interviews  disable trigger interviews_sync_scheduled;",
        "alter table tuyendung.candidates  disable trigger candidates_log;",
        "alter table tuyendung.interviews  disable trigger interviews_log;",
        "alter table tuyendung.onboardings disable trigger onboardings_log;",
        "",
    ]

    cot_uv = [
        "received_at", "full_name", "gender", "region", "phone", "email", "cv_url",
        "position", "department", "level", "source", "screener", "screening_note",
        "status", "hometown", "experience", "available_from", "expected_salary",
        "offer_status", "planned_onboard_date", "actual_onboard_date",
    ]

    L.append(f"insert into tuyendung.candidates ({', '.join(cot_uv)}) values")
    L.append(
        ",\n".join("  (" + ", ".join(q(uv.get(c)) for c in cot_uv) + ")" for uv in ung_vien)
    )
    L.append(";")
    L.append("")

    if phong_van:
        L.append("-- Lịch phỏng vấn: ghép với ứng viên theo số điện thoại, không có thì theo họ tên")
        for pv in phong_van:
            dieu_kien = (
                f"phone_norm = tuyendung.f_phone_norm({q(pv['_sdt'])})"
                if pv["_sdt"]
                else f"full_name = {q(pv['_ten'])}"
            )
            mang = "array[" + ", ".join(q(i) for i in pv["interviewers"]) + "]::text[]"
            L.append(
                "insert into tuyendung.interviews "
                "(candidate_id, round, scheduled_date, scheduled_time, mode, interviewers, result, note, result_email_sent)\n"
                f"select id, {pv['round']}, {q(pv['scheduled_date'])}, {q(pv['scheduled_time'])}, "
                f"{q(pv['mode'])}, {mang}, {q(pv['result'])}, {q(pv['note'])}, {q(pv['result_email_sent'])}\n"
                f"  from tuyendung.candidates where {dieu_kien} "
                "order by created_at desc limit 1\n"
                "on conflict (candidate_id, round) do nothing;"
            )
        L.append("")

    if onboard:
        L.append("-- Onboard: ghép theo họ tên")
        for ob in onboard:
            L.append(
                "insert into tuyendung.onboardings "
                "(candidate_id, onboard_date, office, checklist, assignee_pre, pre_note, "
                "assignee_docs, assignee_training, status, review_7d_result, review_1m_result, "
                "review_2m_result, owner, note)\n"
                f"select id, {q(ob['onboard_date'])}, {q(ob['office'])}, {q(ob['checklist'])}::jsonb, "
                f"{q(ob['assignee_pre'])}, {q(ob['pre_note'])}, {q(ob['assignee_docs'])}, "
                f"{q(ob['assignee_training'])}, {q(ob['status'])}, {q(ob['review_7d_result'])}, "
                f"{q(ob['review_1m_result'])}, {q(ob['review_2m_result'])}, {q(ob['owner'])}, {q(ob['note'])}\n"
                f"  from tuyendung.candidates where full_name = {q(ob['_ten'])} "
                "order by created_at desc limit 1\n"
                "on conflict (candidate_id) do nothing;"
            )
        L.append("")

    L.append("-- Bật lại trigger")
    L.append("alter table tuyendung.interviews  enable trigger interviews_sync_status;")
    L.append("alter table tuyendung.interviews  enable trigger interviews_sync_scheduled;")
    L.append("alter table tuyendung.candidates  enable trigger candidates_log;")
    L.append("alter table tuyendung.interviews  enable trigger interviews_log;")
    L.append("alter table tuyendung.onboardings enable trigger onboardings_log;")
    L.append("")
    L.append("commit;")
    return "\n".join(L)


def day_len_supabase(ung_vien):
    env = doc_env()
    url = env.get("NEXT_PUBLIC_SUPABASE_URL")
    key = env.get("SUPABASE_SERVICE_ROLE_KEY")
    if not (url and key):
        print("Thiếu NEXT_PUBLIC_SUPABASE_URL hoặc SUPABASE_SERVICE_ROLE_KEY trong web/.env.local")
        return 1

    lo = 500
    for i in range(0, len(ung_vien), lo):
        goi = ung_vien[i : i + lo]
        req = urllib.request.Request(
            f"{url}/rest/v1/candidates",
            data=json.dumps(goi, ensure_ascii=False).encode("utf-8"),
            headers={
                "apikey": key,
                "Authorization": f"Bearer {key}",
                "Content-Type": "application/json",
                # bảng nằm ở schema riêng, không phải public
                "Content-Profile": DB_SCHEMA,
                "Accept-Profile": DB_SCHEMA,
                "Prefer": "return=minimal",
            },
            method="POST",
        )
        with urllib.request.urlopen(req) as r:
            print(f"  đã đẩy {i + len(goi)}/{len(ung_vien)} ứng viên (HTTP {r.status})")
    return 0


# ------------------------------------------------------------------ chạy
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--sql", metavar="ĐƯỜNG_DẪN", help="ghi ra file SQL")
    ap.add_argument("--push", action="store_true", help="đẩy thẳng lên Supabase")
    args = ap.parse_args()

    if not XLSX.exists():
        print(f"Không thấy file {XLSX.name}")
        return 1

    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    ung_vien, phong_van = doc_ung_vien(wb)
    onboard = doc_onboard(wb)

    print(f"Đọc từ {XLSX.name}:")
    print(f"  ứng viên          {len(ung_vien)}")
    print(f"  lịch phỏng vấn    {len(phong_van)}")
    print(f"  dòng onboard      {len(onboard)}")

    thieu_sdt = sum(1 for u in ung_vien if not u["phone"])
    thieu_mail = sum(1 for u in ung_vien if not u["email"])
    print(f"\nChất lượng dữ liệu:")
    print(f"  thiếu số điện thoại  {thieu_sdt}")
    print(f"  thiếu email          {thieu_mail}")

    trung = defaultdict(list)
    for u in ung_vien:
        if u["phone"]:
            trung[u["phone"]].append(u["full_name"])
    nghi_trung = {k: v for k, v in trung.items() if len(v) > 1}
    print(f"  nghi trùng theo SĐT  {len(nghi_trung)}")
    for k, v in list(nghi_trung.items())[:5]:
        print(f"     {k}: {', '.join(v[:4])}")

    nguon = Counter(u["source"] for u in ung_vien if u["source"])
    if nguon:
        print("\n  nguồn CV sau khi gộp:", ", ".join(f"{k}={v}" for k, v in nguon.most_common(8)))

    if args.sql:
        out = pathlib.Path(args.sql)
        if not out.is_absolute():
            out = ROOT / out
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(sinh_sql(ung_vien, phong_van, onboard), encoding="utf-8")
        print(f"\nĐã ghi SQL: {out.relative_to(ROOT)}")

    if args.push:
        return day_len_supabase(ung_vien)

    if not args.sql and not args.push:
        print("\n(Chạy thử, chưa ghi gì. Thêm --sql <file> hoặc --push để nạp thật.)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
